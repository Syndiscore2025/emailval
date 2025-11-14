"""
Reporting Module
Handles export and report generation for email validation results
"""
import csv
import io
from typing import List, Dict, Any
from datetime import datetime


def generate_csv_report(validation_results: List[Dict[str, Any]]) -> str:
    """
    Generate CSV report from validation results.
    
    Args:
        validation_results: List of validation result dictionaries
        
    Returns:
        CSV content as string
    """
    output = io.StringIO()
    
    # Define CSV headers
    headers = ['Email', 'Status', 'Email Type', 'Is Disposable', 'Is Role Based', 
               'Has MX Records', 'SMTP Valid', 'Errors', 'Validation Date']
    
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    
    for result in validation_results:
        email = result.get('email', '')
        valid = result.get('valid', False)
        checks = result.get('checks', {})
        
        # Extract type info
        type_check = checks.get('type', {})
        email_type = type_check.get('email_type', 'unknown')
        is_disposable = type_check.get('is_disposable', False)
        is_role_based = type_check.get('is_role_based', False)
        
        # Extract domain info
        domain_check = checks.get('domain', {})
        has_mx = domain_check.get('has_mx', False)
        
        # Extract SMTP info
        smtp_check = checks.get('smtp', {})
        smtp_valid = smtp_check.get('valid', False)
        
        # Get errors
        errors = '; '.join(result.get('errors', []))
        
        writer.writerow({
            'Email': email,
            'Status': 'Valid' if valid else 'Invalid',
            'Email Type': email_type,
            'Is Disposable': 'Yes' if is_disposable else 'No',
            'Is Role Based': 'Yes' if is_role_based else 'No',
            'Has MX Records': 'Yes' if has_mx else 'No',
            'SMTP Valid': 'Yes' if smtp_valid else 'No',
            'Errors': errors,
            'Validation Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return output.getvalue()


def generate_excel_report(validation_results: List[Dict[str, Any]]) -> bytes:
    """
    Generate Excel report from validation results.
    
    Args:
        validation_results: List of validation result dictionaries
        
    Returns:
        Excel file content as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Results"
        
        # Define headers
        headers = ['Email', 'Status', 'Email Type', 'Is Disposable', 'Is Role Based', 
                   'Has MX Records', 'SMTP Valid', 'Errors', 'Validation Date']
        
        # Write headers with styling
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, result in enumerate(validation_results, start=2):
            email = result.get('email', '')
            valid = result.get('valid', False)
            checks = result.get('checks', {})
            
            # Extract type info
            type_check = checks.get('type', {})
            email_type = type_check.get('email_type', 'unknown')
            is_disposable = type_check.get('is_disposable', False)
            is_role_based = type_check.get('is_role_based', False)
            
            # Extract domain info
            domain_check = checks.get('domain', {})
            has_mx = domain_check.get('has_mx', False)
            
            # Extract SMTP info
            smtp_check = checks.get('smtp', {})
            smtp_valid = smtp_check.get('valid', False)
            
            # Get errors
            errors = '; '.join(result.get('errors', []))
            
            ws.cell(row=row_idx, column=1, value=email)
            ws.cell(row=row_idx, column=2, value='Valid' if valid else 'Invalid')
            ws.cell(row=row_idx, column=3, value=email_type)
            ws.cell(row=row_idx, column=4, value='Yes' if is_disposable else 'No')
            ws.cell(row=row_idx, column=5, value='Yes' if is_role_based else 'No')
            ws.cell(row=row_idx, column=6, value='Yes' if has_mx else 'No')
            ws.cell(row=row_idx, column=7, value='Yes' if smtp_valid else 'No')
            ws.cell(row=row_idx, column=8, value=errors)
            ws.cell(row=row_idx, column=9, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


def generate_pdf_report(validation_results: List[Dict[str, Any]], summary_stats: Dict[str, Any] = None) -> bytes:
    """
    Generate PDF report from validation results.

    Args:
        validation_results: List of validation result dictionaries
        summary_stats: Optional summary statistics

    Returns:
        PDF file content as bytes
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        title = Paragraph("Email Validation Report", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2*inch))

        # Date
        date_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_para = Paragraph(date_text, styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 0.3*inch))

        # Summary Statistics
        if summary_stats:
            summary_heading = Paragraph("Summary Statistics", heading_style)
            story.append(summary_heading)

            summary_data = [
                ['Metric', 'Value'],
                ['Total Emails', str(summary_stats.get('total', len(validation_results)))],
                ['Valid Emails', str(summary_stats.get('valid', 0))],
                ['Invalid Emails', str(summary_stats.get('invalid', 0))],
                ['Valid Percentage', f"{summary_stats.get('valid_percent', 0):.1f}%"],
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))

        # Validation Results
        results_heading = Paragraph("Detailed Validation Results", heading_style)
        story.append(results_heading)

        # Limit to first 100 results for PDF
        limited_results = validation_results[:100]

        # Create table data
        table_data = [['Email', 'Status', 'Type', 'Errors']]

        for result in limited_results:
            email = result.get('email', '')
            valid = result.get('valid', False)
            checks = result.get('checks', {})
            type_check = checks.get('type', {})
            email_type = type_check.get('email_type', 'unknown')
            errors = ', '.join(result.get('errors', []))[:50]  # Truncate long errors

            # Truncate long emails
            if len(email) > 30:
                email = email[:27] + '...'

            table_data.append([
                email,
                'Valid' if valid else 'Invalid',
                email_type,
                errors if errors else 'None'
            ])

        # Create table
        results_table = Table(table_data, colWidths=[2.2*inch, 1*inch, 1.3*inch, 2*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(results_table)

        # Note if results were truncated
        if len(validation_results) > 100:
            story.append(Spacer(1, 0.2*inch))
            note = Paragraph(
                f"<i>Note: Showing first 100 of {len(validation_results)} results. "
                "Download CSV or Excel for complete data.</i>",
                styles['Normal']
            )
            story.append(note)

        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.getvalue()

    except ImportError as e:
        raise ImportError(f"reportlab is required for PDF export. Install with: pip install reportlab. Error: {e}")

